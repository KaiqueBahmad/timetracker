import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from services.companyService import get_company_id, get_time_records

def exportToExcel(empresa, date_obj=None):
    try:
        empresa_id = get_company_id(empresa)
        if not empresa_id:
            print(f"Empresa '{empresa}' não encontrada")
            return False
        
        records = get_time_records(empresa_id, date_obj)

        wb = Workbook()
        ws = wb.active
        ws.title = "Registro de Horas"
        
        if date_obj:
            month_year = date_obj.strftime('%B %Y')
            file_date = date_obj.strftime('%m_%Y')
        else:
            now = datetime.now()
            month_year = now.strftime('%B %Y')
            file_date = now.strftime('%m_%Y')
        
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Relatório de Horas - {empresa} - {month_year}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        headers = ['ID', 'Data Início', 'Hora Início', 'Data Fim', 'Hora Fim', 'Duração (min)']
        header_row = 3
        
        header_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f'{col_letter}{header_row}']
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            
            ws.column_dimensions[col_letter].width = max(len(header) + 4, 12)
        
        row_num = header_row + 1
        for record in records:
            record_id, inicio, fim, duracao = record
            
            inicio_dt = datetime.fromisoformat(inicio.replace(' ', 'T'))
            fim_dt = datetime.fromisoformat(fim.replace(' ', 'T')) if fim else None
            
            inicio_data = inicio_dt.strftime('%Y-%m-%d')
            inicio_hora = inicio_dt.strftime('%H:%M:%S')
            
            if fim_dt:
                fim_data = fim_dt.strftime('%Y-%m-%d')
                fim_hora = fim_dt.strftime('%H:%M:%S')
            else:
                fim_data = ''
                fim_hora = ''
            
            row_data = [record_id, inicio_data, inicio_hora, fim_data, fim_hora, duracao]
            for col_num, value in enumerate(row_data, 1):
                cell = ws[f'{get_column_letter(col_num)}{row_num}']
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            
            row_num += 1
        
        total_row = row_num + 1
        ws[f'A{total_row}'] = "Total de Horas"
        ws[f'A{total_row}'].font = Font(bold=True)
        
        total_minutes = sum(record[3] or 0 for record in records)
        total_hours = total_minutes / 60
        
        hours = int(total_hours)
        minutes = int((total_hours - hours) * 60)
        ws[f'F{total_row}'] = f"{hours}h {minutes}min ({total_minutes} min)"
        ws[f'F{total_row}'].font = Font(bold=True)
        
        os.makedirs("exports", exist_ok=True)
        
        filename = f"{empresa}_{file_date}.xlsx"
        filepath = os.path.join("exports", filename)
        wb.save(filepath)
        
        print(f"Relatório Excel exportado com sucesso para: {filepath}")
        return True
        
    except Exception as e:
        print(f"Erro ao exportar para Excel: {e}")
        return False